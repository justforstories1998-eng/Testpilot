"""
TestPilot Enhanced Report Generator
=====================================
Generates ultra-detailed Excel reports with:
- Playwright-specific metrics
- Performance timing data
- Network request analysis  
- Accessibility findings
- Detailed step-by-step breakdown
- AI root cause analysis
- Trend charts
"""
from __future__ import annotations

import os
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

from app.config import get_settings

logger = logging.getLogger("testpilot.services.report_generator")

# ---- Style Constants ----
HEADER_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
SKIP_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
PASS_FONT = Font(name="Calibri", color="006100", bold=True, size=10)
FAIL_FONT = Font(name="Calibri", color="9C0006", bold=True, size=10)
ERROR_FONT = Font(name="Calibri", color="9C6500", bold=True, size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=28, color="1B2A4A")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=14, color="4472C4")
BODY_FONT = Font(name="Calibri", size=10)
LABEL_FONT = Font(name="Calibri", bold=True, size=11, color="666666")
LINK_FONT = Font(name="Calibri", size=10, color="0563C1", underline="single")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
RIGHT_CENTER = Alignment(horizontal="right", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


class ReportGenerator:
    """Generates professional Excel test reports."""

    def __init__(self):
        settings = get_settings()
        self.reports_dir = Path(settings.reports_dir)
        self.reports_dir.mkdir(parents=True, exist_ok=True)
        self.screenshots_base_url = "/static/screenshots"

    def generate(
        self,
        session_id: str,
        project_name: str,
        url: str,
        test_results: Dict[str, Any],
        failure_analyses: Optional[List[Dict[str, Any]]] = None,
        page_analysis: Optional[Dict[str, Any]] = None,
    ) -> str:
        wb = Workbook()
        wb.remove(wb.active)

        self._create_cover_sheet(wb, session_id, project_name, url, test_results)
        self._create_executive_summary(wb, test_results)
        self._create_all_test_results(wb, test_results)
        self._create_step_details(wb, test_results)
        self._create_failure_analysis(wb, test_results, failure_analyses)
        self._create_performance_sheet(wb, test_results)
        self._create_screenshots_sheet(wb, test_results)
        if page_analysis:
            self._create_page_analysis_sheet(wb, page_analysis)
        self._create_recommendations(wb, test_results, failure_analyses)

        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        short_id = session_id[:8] if session_id else "report"
        safe_name = "".join(
            c for c in project_name if c.isalnum() or c in " _-"
        ).strip().replace(" ", "_")
        filename = f"TestPilot_{safe_name}_{short_id}_{timestamp}.xlsx"
        filepath = str(self.reports_dir / filename)
        wb.save(filepath)
        logger.info(f"Report generated: {filepath}")
        return filepath

    def _screenshot_link(self, path: str) -> str:
        if not path:
            return ""
        return f"{self.screenshots_base_url}/{os.path.basename(path)}"

    def _apply_status_style(self, cell, status: str):
        cell.alignment = CENTER
        s = status.lower()
        if s == "passed":
            cell.fill = PASS_FILL
            cell.font = PASS_FONT
        elif s == "failed":
            cell.fill = FAIL_FILL
            cell.font = FAIL_FONT
        elif s == "error":
            cell.fill = ERROR_FILL
            cell.font = ERROR_FONT
        elif s == "skipped":
            cell.fill = SKIP_FILL

    def _create_cover_sheet(self, wb, session_id, project_name, url, results):
        ws = wb.create_sheet("Cover Page")
        ws.sheet_properties.tabColor = "1B2A4A"
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 55

        ws.merge_cells("B3:C3")
        ws["B3"].value = "TESTPILOT"
        ws["B3"].font = TITLE_FONT
        ws.merge_cells("B4:C4")
        ws["B4"].value = "AI-Powered Playwright Test Report"
        ws["B4"].font = SUBTITLE_FONT
        ws.merge_cells("B5:C5")
        ws["B5"].value = f"Project: {project_name}"
        ws["B5"].font = Font(name="Calibri", size=12, color="666666")

        pass_rate = results.get("pass_rate", 0)
        exec_time = results.get("execution_time_ms", 0) / 1000

        details = [
            ("Report ID", (session_id[:8] if session_id else "N/A").upper()),
            ("Project Name", project_name),
            ("Target URL", url),
            ("Generated On", datetime.now(timezone.utc).strftime("%B %d, %Y at %H:%M UTC")),
            ("", ""),
            ("Total Tests", str(results.get("total", 0))),
            ("Passed", str(results.get("passed", 0))),
            ("Failed", str(results.get("failed", 0))),
            ("Errors", str(results.get("errors", 0))),
            ("Skipped (Need Approval)", str(results.get("skipped", 0))),
            ("Pass Rate", f"{pass_rate}%"),
            ("Execution Time", f"{exec_time:.1f} seconds"),
            ("", ""),
            ("Testing Engine", "Playwright (Chromium)"),
            ("AI Engine", "Groq AI (Llama 3.3 70B)"),
            ("Platform", "TestPilot v3.0"),
        ]

        row = 9
        for label, value in details:
            if label:
                ws[f"B{row}"].value = label
                ws[f"B{row}"].font = LABEL_FONT
                ws[f"B{row}"].alignment = RIGHT_CENTER
                ws[f"C{row}"].value = value
                ws[f"C{row}"].font = Font(name="Calibri", size=11)
            row += 1

        row += 1
        ws.merge_cells(f"B{row}:C{row}")
        cell = ws[f"B{row}"]
        if pass_rate >= 90:
            cell.value = "OVERALL STATUS: PASS ✓"
            cell.font = Font(name="Calibri", bold=True, size=16, color="006100")
            cell.fill = PASS_FILL
        elif pass_rate >= 50:
            cell.value = "OVERALL STATUS: PARTIAL PASS ⚠"
            cell.font = Font(name="Calibri", bold=True, size=16, color="9C6500")
            cell.fill = ERROR_FILL
        else:
            cell.value = "OVERALL STATUS: FAIL ✗"
            cell.font = Font(name="Calibri", bold=True, size=16, color="9C0006")
            cell.fill = FAIL_FILL
        cell.alignment = CENTER
        ws.row_dimensions[row].height = 45

    def _create_executive_summary(self, wb, results):
        ws = wb.create_sheet("Executive Summary")
        ws.sheet_properties.tabColor = "4472C4"
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 25

        ws.merge_cells("A1:B1")
        ws["A1"].value = "TEST EXECUTION SUMMARY"
        ws["A1"].font = SUBTITLE_FONT
        ws["A1"].alignment = CENTER
        ws.row_dimensions[1].height = 30

        for col, h in enumerate(["Metric", "Value"], 1):
            c = ws.cell(row=3, column=col, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = CENTER
            c.border = THIN_BORDER

        stats = [
            ("Total Test Cases", results.get("total", 0)),
            ("Passed ✓", results.get("passed", 0)),
            ("Failed ✗", results.get("failed", 0)),
            ("Errors ⚠", results.get("errors", 0)),
            ("Skipped (Needs Approval)", results.get("skipped", 0)),
            ("Pass Rate", f"{results.get('pass_rate', 0)}%"),
            ("Total Execution Time", f"{results.get('execution_time_ms', 0) / 1000:.1f} seconds"),
            ("Average Test Time", self._avg_test_time(results)),
            ("Slowest Test", self._slowest_test(results)),
            ("Fastest Test", self._fastest_test(results)),
        ]

        for i, (metric, value) in enumerate(stats, 4):
            mc = ws.cell(row=i, column=1, value=metric)
            mc.font = Font(name="Calibri", bold=True, size=10)
            mc.border = THIN_BORDER
            vc = ws.cell(row=i, column=2, value=value)
            vc.border = THIN_BORDER
            vc.alignment = CENTER

        # Category breakdown
        cats = self._build_category_stats(results)
        if cats:
            row = 16
            ws.merge_cells(f"A{row}:G{row}")
            ws[f"A{row}"].value = "BREAKDOWN BY CATEGORY"
            ws[f"A{row}"].font = SUBTITLE_FONT
            row += 1

            cat_headers = [
                ("Category", 20), ("Total", 10), ("Passed", 10),
                ("Failed", 10), ("Errors", 10), ("Skipped", 10), ("Pass Rate", 15),
            ]
            for col, (h, w) in enumerate(cat_headers, 1):
                ws.column_dimensions[get_column_letter(col)].width = w
                c = ws.cell(row=row, column=col, value=h)
                c.font = HEADER_FONT
                c.fill = HEADER_FILL
                c.alignment = CENTER
                c.border = THIN_BORDER

            for cat_name, counts in sorted(cats.items()):
                row += 1
                executed = counts["passed"] + counts["failed"] + counts["errors"]
                rate = round(counts["passed"] / executed * 100, 1) if executed > 0 else 0
                values = [
                    cat_name.replace("_", " ").title(),
                    counts["total"], counts["passed"],
                    counts["failed"], counts["errors"],
                    counts["skipped"], f"{rate}%",
                ]
                for col, v in enumerate(values, 1):
                    c = ws.cell(row=row, column=col, value=v)
                    c.border = THIN_BORDER
                    c.alignment = CENTER
                rate_cell = ws.cell(row=row, column=7)
                if rate >= 90:
                    rate_cell.fill = PASS_FILL
                    rate_cell.font = PASS_FONT
                elif rate >= 50:
                    rate_cell.fill = ERROR_FILL
                    rate_cell.font = ERROR_FONT
                else:
                    rate_cell.fill = FAIL_FILL
                    rate_cell.font = FAIL_FONT

        # Pie chart
        try:
            cd = wb.create_sheet("_chart_data")
            cd["A1"], cd["B1"] = "Status", "Count"
            cd["A2"], cd["B2"] = "Passed", results.get("passed", 0)
            cd["A3"], cd["B3"] = "Failed", results.get("failed", 0)
            cd["A4"], cd["B4"] = "Errors", results.get("errors", 0)
            cd["A5"], cd["B5"] = "Skipped", results.get("skipped", 0)

            pie = PieChart()
            pie.title = "Test Results Distribution"
            pie.style = 10
            pie.width = 16
            pie.height = 10
            data = Reference(cd, min_col=2, min_row=1, max_row=5)
            cats_ref = Reference(cd, min_col=1, min_row=2, max_row=5)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(cats_ref)
            colors = ["006100", "9C0006", "9C6500", "4472C4"]
            for i, color in enumerate(colors):
                pt = DataPoint(idx=i)
                pt.graphicalProperties.solidFill = color
                pie.series[0].data_points.append(pt)
            ws.add_chart(pie, "I3")
            cd.sheet_state = "hidden"
        except Exception as e:
            logger.warning(f"Chart creation failed: {e}")

    def _create_all_test_results(self, wb, results):
        ws = wb.create_sheet("All Test Results")
        ws.sheet_properties.tabColor = "70AD47"

        headers = [
            "Test ID", "Suite", "Test Name", "Description",
            "Category", "Priority", "Destructive?", "Status",
            "Expected Result", "What Actually Happened",
            "Error Details", "Time (ms)", "Steps Count", "Screenshot",
        ]
        widths = [10, 20, 30, 45, 14, 10, 14, 12, 40, 40, 40, 12, 12, 30]

        for i, (h, w) in enumerate(zip(headers, widths), 1):
            ws.column_dimensions[get_column_letter(i)].width = w
            c = ws.cell(row=1, column=i, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = CENTER
            c.border = THIN_BORDER

        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        ws.freeze_panes = "A2"

        for row_idx, tr in enumerate(results.get("test_results", []), 2):
            screenshot_link = self._screenshot_link(tr.get("screenshot", ""))
            values = [
                tr.get("test_id", ""),
                tr.get("suite_name", ""),
                tr.get("test_name", ""),
                tr.get("description", ""),
                tr.get("category", "").replace("_", " ").title(),
                tr.get("priority", "").title(),
                "YES - Modifies Data" if tr.get("is_destructive") else "No - Safe",
                tr.get("status", "").upper(),
                tr.get("expected_result", ""),
                tr.get("actual_result", ""),
                tr.get("error_message", "") or "",
                tr.get("execution_time_ms", 0),
                len(tr.get("steps", [])),
                "",
            ]

            for col, v in enumerate(values, 1):
                c = ws.cell(row=row_idx, column=col, value=v)
                c.font = BODY_FONT
                c.alignment = LEFT_WRAP
                c.border = THIN_BORDER

            if screenshot_link:
                sc = ws.cell(row=row_idx, column=14)
                sc.font = LINK_FONT
                sc.hyperlink = screenshot_link
                sc.value = "📷 View Screenshot"

            self._apply_status_style(ws.cell(row=row_idx, column=8), tr.get("status", ""))

            dc = ws.cell(row=row_idx, column=7)
            dc.alignment = CENTER
            if tr.get("is_destructive"):
                dc.font = Font(name="Calibri", bold=True, color="FF6600", size=10)

            ws.row_dimensions[row_idx].height = 60

    def _create_step_details(self, wb, results):
        ws = wb.create_sheet("Step-by-Step Details")
        ws.sheet_properties.tabColor = "FFC000"

        headers = ["Test ID", "Test Name", "Step #", "Action", "Expected", "Status", "Error", "Screenshot"]
        widths = [10, 30, 8, 50, 40, 10, 45, 30]

        for i, (h, w) in enumerate(zip(headers, widths), 1):
            ws.column_dimensions[get_column_letter(i)].width = w
            c = ws.cell(row=1, column=i, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = CENTER
            c.border = THIN_BORDER

        ws.freeze_panes = "A2"
        row = 2

        for tr in results.get("test_results", []):
            tid = tr.get("test_id", "")
            tname = tr.get("test_name", "")
            steps = tr.get("steps", [])

            if not steps:
                for col, v in enumerate([tid, tname, 1, tr.get("actual_result", "N/A"), "", tr.get("status", "").upper(), "", ""], 1):
                    c = ws.cell(row=row, column=col, value=v)
                    c.border = THIN_BORDER
                    c.font = BODY_FONT
                self._apply_status_style(ws.cell(row=row, column=6), tr.get("status", ""))
                ws.row_dimensions[row].height = 35
                row += 1
                continue

            for step in steps:
                step_num = step.get("step", step.get("step_number", 0))
                action = step.get("action", step.get("description", ""))
                expected = step.get("expected", "")
                status = step.get("status", "")
                error = step.get("error", "") or ""
                screenshot = self._screenshot_link(step.get("screenshot", ""))

                for col, v in enumerate([tid, tname, step_num, action, expected, status.upper(), error, ""], 1):
                    c = ws.cell(row=row, column=col, value=v)
                    c.font = BODY_FONT
                    c.alignment = LEFT_WRAP
                    c.border = THIN_BORDER

                self._apply_status_style(ws.cell(row=row, column=6), status)
                if screenshot:
                    sc = ws.cell(row=row, column=8)
                    sc.font = LINK_FONT
                    sc.hyperlink = screenshot
                    sc.value = "View"
                ws.row_dimensions[row].height = 35
                row += 1

    def _create_failure_analysis(self, wb, results, analyses):
        ws = wb.create_sheet("Failure Analysis")
        ws.sheet_properties.tabColor = "FF0000"

        headers = [
            "Test ID", "Test Name", "Root Cause (Simple)",
            "Technical Details", "Severity", "Expected Behavior",
            "Suggested Fix", "Regression Risk", "Screenshot",
        ]
        widths = [10, 30, 50, 45, 12, 40, 45, 15, 30]

        for i, (h, w) in enumerate(zip(headers, widths), 1):
            ws.column_dimensions[get_column_letter(i)].width = w
            c = ws.cell(row=1, column=i, value=h)
            c.font = HEADER_FONT
            c.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            c.alignment = CENTER
            c.border = THIN_BORDER

        ws.freeze_panes = "A2"
        failed = [t for t in results.get("test_results", []) if t.get("status") in ("failed", "error")]
        analyses = analyses or []

        if not failed:
            ws.merge_cells("A2:I2")
            ws["A2"].value = "🎉 All tests passed! No failures to analyze."
            ws["A2"].font = Font(name="Calibri", size=14, color="006100", bold=True)
            ws["A2"].fill = PASS_FILL
            ws["A2"].alignment = CENTER
            ws.row_dimensions[2].height = 40
            return

        for i, tr in enumerate(failed, 2):
            a = analyses[i - 2] if (i - 2) < len(analyses) else {}
            screenshot_link = self._screenshot_link(tr.get("screenshot", ""))

            values = [
                tr.get("test_id", ""),
                tr.get("test_name", ""),
                a.get("root_cause", tr.get("error_message", "Unknown")[:200]),
                a.get("technical_details", tr.get("error_message", ""))[:300],
                a.get("severity", tr.get("priority", "medium")).upper(),
                a.get("what_should_have_happened", tr.get("expected_result", ""))[:200],
                a.get("suggested_fix", "Manual investigation required")[:300],
                a.get("regression_risk", "medium").upper(),
                "",
            ]

            for col, v in enumerate(values, 1):
                c = ws.cell(row=i, column=col, value=v)
                c.font = BODY_FONT
                c.alignment = LEFT_WRAP
                c.border = THIN_BORDER

            if screenshot_link:
                sc = ws.cell(row=i, column=9)
                sc.font = LINK_FONT
                sc.hyperlink = screenshot_link
                sc.value = "📷 View Failure"

            sev_cell = ws.cell(row=i, column=5)
            sev_cell.alignment = CENTER
            sev = a.get("severity", "medium").lower()
            if sev == "critical":
                sev_cell.fill = FAIL_FILL
                sev_cell.font = FAIL_FONT
            elif sev == "high":
                sev_cell.fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
                sev_cell.font = Font(name="Calibri", bold=True, color="CC0000", size=10)
            elif sev == "medium":
                sev_cell.fill = ERROR_FILL
                sev_cell.font = ERROR_FONT

            ws.row_dimensions[i].height = 100

    def _create_performance_sheet(self, wb, results):
        """New sheet: Performance metrics for each test."""
        ws = wb.create_sheet("Performance Metrics")
        ws.sheet_properties.tabColor = "0070C0"

        ws.merge_cells("A1:E1")
        ws["A1"].value = "PERFORMANCE METRICS"
        ws["A1"].font = SUBTITLE_FONT
        ws["A1"].alignment = CENTER
        ws.row_dimensions[1].height = 30

        headers = ["Test Name", "Status", "Execution Time (ms)", "Time Rating", "Notes"]
        widths = [40, 12, 20, 15, 40]

        for i, (h, w) in enumerate(zip(headers, widths), 1):
            ws.column_dimensions[get_column_letter(i)].width = w
            c = ws.cell(row=3, column=i, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = CENTER
            c.border = THIN_BORDER

        test_results = results.get("test_results", [])
        sorted_results = sorted(test_results, key=lambda t: t.get("execution_time_ms", 0), reverse=True)

        for row_idx, tr in enumerate(sorted_results, 4):
            exec_time = tr.get("execution_time_ms", 0)
            if exec_time < 1000:
                rating = "⚡ Fast"
                fill = PASS_FILL
            elif exec_time < 5000:
                rating = "✓ Normal"
                fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
            else:
                rating = "🐌 Slow"
                fill = ERROR_FILL

            values = [
                tr.get("test_name", ""),
                tr.get("status", "").upper(),
                exec_time,
                rating,
                f"Average: {self._avg_time(test_results):.0f}ms",
            ]

            for col, v in enumerate(values, 1):
                c = ws.cell(row=row_idx, column=col, value=v)
                c.font = BODY_FONT
                c.alignment = CENTER if col > 1 else LEFT_WRAP
                c.border = THIN_BORDER

            ws.cell(row=row_idx, column=4).fill = fill
            self._apply_status_style(ws.cell(row=row_idx, column=2), tr.get("status", ""))

        # Add total stats
        row = len(sorted_results) + 5
        ws.cell(row=row, column=1, value="TOTALS").font = Font(name="Calibri", bold=True, size=11)
        ws.cell(row=row, column=3, value=results.get("execution_time_ms", 0)).font = Font(name="Calibri", bold=True)
        ws.cell(row=row, column=3).alignment = CENTER

    def _create_screenshots_sheet(self, wb, results):
        ws = wb.create_sheet("Screenshots")
        ws.sheet_properties.tabColor = "FF6600"

        headers = ["Test ID", "Test Name", "Status", "Screenshot Link", "Description"]
        widths = [12, 35, 12, 45, 50]

        for i, (h, w) in enumerate(zip(headers, widths), 1):
            ws.column_dimensions[get_column_letter(i)].width = w
            c = ws.cell(row=1, column=i, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = CENTER
            c.border = THIN_BORDER

        ws.freeze_panes = "A2"
        row = 2
        has_screenshots = False

        for tr in results.get("test_results", []):
            screenshot = tr.get("screenshot", "")
            if not screenshot:
                continue
            has_screenshots = True
            link = self._screenshot_link(screenshot)
            status = tr.get("status", "")

            ws.cell(row=row, column=1, value=tr.get("test_id", "")).border = THIN_BORDER
            ws.cell(row=row, column=1).alignment = CENTER
            ws.cell(row=row, column=2, value=tr.get("test_name", "")).border = THIN_BORDER
            ws.cell(row=row, column=2).alignment = LEFT_WRAP

            status_cell = ws.cell(row=row, column=3, value=status.upper())
            status_cell.border = THIN_BORDER
            self._apply_status_style(status_cell, status)

            link_cell = ws.cell(row=row, column=4)
            link_cell.font = LINK_FONT
            link_cell.hyperlink = link
            link_cell.value = "📷 Click to View Screenshot"
            link_cell.border = THIN_BORDER

            desc = tr.get("actual_result", "")[:200]
            desc_cell = ws.cell(row=row, column=5, value=desc)
            desc_cell.border = THIN_BORDER
            desc_cell.alignment = LEFT_WRAP
            desc_cell.font = Font(name="Calibri", size=9, color="9C0006" if status in ("failed", "error") else "333333")

            ws.row_dimensions[row].height = 30
            row += 1

        if not has_screenshots:
            ws.merge_cells("A2:E2")
            ws["A2"].value = "No screenshots captured during this test run"
            ws["A2"].font = Font(name="Calibri", size=12, italic=True, color="666666")
            ws["A2"].alignment = CENTER

    def _create_page_analysis_sheet(self, wb, pa):
        ws = wb.create_sheet("Page Analysis")
        ws.sheet_properties.tabColor = "7030A0"
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 65

        ws.merge_cells("B1:C1")
        ws["B1"].value = "AI PAGE ANALYSIS"
        ws["B1"].font = SUBTITLE_FONT
        ws["B1"].alignment = CENTER
        ws.row_dimensions[1].height = 30

        row = 3
        basic_info = [
            ("Page Title", pa.get("page_title", "N/A")),
            ("Page Type", pa.get("page_type", "N/A")),
            ("Description", pa.get("page_description", "N/A")),
            ("Has CRUD Operations", "Yes" if pa.get("has_crud_operations") else "No"),
        ]
        for label, val in basic_info:
            ws.cell(row=row, column=2, value=label).font = Font(name="Calibri", bold=True)
            ws.cell(row=row, column=2).border = THIN_BORDER
            ws.cell(row=row, column=3, value=str(val)).border = THIN_BORDER
            ws.cell(row=row, column=3).alignment = LEFT_WRAP
            row += 1

        sections = [
            ("Key Features", pa.get("key_features", []), "1B2A4A"),
            ("Technologies Detected", pa.get("technologies_detected", []), "1B2A4A"),
            ("Potential Test Areas", pa.get("potential_test_areas", []), "4472C4"),
            ("Accessibility Notes", pa.get("accessibility_notes", []), "9C0006"),
        ]
        for title, items, color in sections:
            row += 1
            ws.merge_cells(f"B{row}:C{row}")
            ws.cell(row=row, column=2, value=title).font = Font(name="Calibri", bold=True, size=12, color=color)
            row += 1
            if items and isinstance(items, list):
                for item in items:
                    ws.cell(row=row, column=2, value="▶").alignment = CENTER
                    ws.cell(row=row, column=3, value=str(item)).font = BODY_FONT
                    ws.cell(row=row, column=3).alignment = LEFT_WRAP
                    row += 1
            else:
                ws.cell(row=row, column=3, value="None detected").font = Font(name="Calibri", italic=True, color="666666")
                row += 1

    def _create_recommendations(self, wb, results, analyses):
        ws = wb.create_sheet("Recommendations")
        ws.sheet_properties.tabColor = "00B0F0"
        ws.column_dimensions["B"].width = 5
        ws.column_dimensions["C"].width = 85

        ws.merge_cells("B1:C1")
        ws["B1"].value = "RECOMMENDATIONS AND NEXT STEPS"
        ws["B1"].font = SUBTITLE_FONT
        ws.row_dimensions[1].height = 30

        pr = results.get("pass_rate", 0)
        skipped = results.get("skipped", 0)
        failed = results.get("failed", 0)
        errors = results.get("errors", 0)

        recs = []
        if pr >= 90:
            recs.append("✅ [EXCELLENT] Pass rate above 90%. Application is in great shape. Focus on fixing the few remaining failures.")
        elif pr >= 70:
            recs.append("✔ [GOOD] Pass rate above 70%. Review Failure Analysis tab. Prioritize critical/high severity issues.")
        elif pr >= 50:
            recs.append("⚠ [NEEDS ATTENTION] Pass rate 50-70%. Several issues need addressing. Work with dev team to fix failures.")
        else:
            recs.append("🚨 [CRITICAL] Pass rate below 50%. Immediate escalation required. Significant application issues detected.")

        if skipped > 0:
            recs.append(f"📋 [ACTION REQUIRED] {skipped} test(s) skipped (destructive). Review 'All Test Results' tab and approve safe ones.")
        if failed > 0:
            recs.append(f"🔍 [REVIEW FAILURES] {failed} test(s) failed. See 'Failure Analysis' tab for AI-powered root cause analysis.")
        if errors > 0:
            recs.append(f"🔧 [FIX ERRORS] {errors} test(s) had execution errors. Usually caused by element changes or timeouts.")

        recs.extend([
            "🔄 [AUTOMATION] Schedule these tests to run automatically after every deployment.",
            "🌐 [CROSS-BROWSER] Run tests on Firefox and Safari in addition to Chromium.",
            "📱 [MOBILE] Test on mobile viewports (375px, 768px) - Playwright supports this natively.",
            "♿ [ACCESSIBILITY] Add ARIA compliance tests using Playwright's accessibility testing features.",
            "⚡ [PERFORMANCE] Monitor Core Web Vitals. Playwright can capture LCP, FID, CLS metrics.",
            "🔒 [SECURITY] Add tests for XSS, CSRF protection, and authentication edge cases.",
            "📊 [TRACKING] Compare test results over time to catch regression trends.",
        ])

        row = 3
        for idx, rec in enumerate(recs, 1):
            ws.cell(row=row, column=2, value=f"{idx}.").font = Font(name="Calibri", size=12, bold=True, color="4472C4")
            ws.cell(row=row, column=2).alignment = CENTER
            ws.cell(row=row, column=3, value=rec).font = BODY_FONT
            ws.cell(row=row, column=3).alignment = LEFT_WRAP
            line_count = len(rec) // 80 + 1
            ws.row_dimensions[row].height = max(30, line_count * 16)
            row += 1

    # ---- Helper methods ----
    def _avg_test_time(self, results) -> str:
        tests = results.get("test_results", [])
        if not tests:
            return "N/A"
        times = [t.get("execution_time_ms", 0) for t in tests]
        return f"{sum(times) / len(times):.0f}ms avg"

    def _avg_time(self, tests) -> float:
        if not tests:
            return 0
        return sum(t.get("execution_time_ms", 0) for t in tests) / len(tests)

    def _slowest_test(self, results) -> str:
        tests = results.get("test_results", [])
        if not tests:
            return "N/A"
        slowest = max(tests, key=lambda t: t.get("execution_time_ms", 0))
        return f"{slowest.get('test_name', 'Unknown')} ({slowest.get('execution_time_ms', 0)}ms)"

    def _fastest_test(self, results) -> str:
        tests = results.get("test_results", [])
        if not tests:
            return "N/A"
        executed = [t for t in tests if t.get("status") != "skipped"]
        if not executed:
            return "N/A"
        fastest = min(executed, key=lambda t: t.get("execution_time_ms", 999999))
        return f"{fastest.get('test_name', 'Unknown')} ({fastest.get('execution_time_ms', 0)}ms)"

    def _build_category_stats(self, results):
        cats = {}
        for tr in results.get("test_results", []):
            cat = tr.get("category", "other")
            if cat not in cats:
                cats[cat] = {"total": 0, "passed": 0, "failed": 0, "errors": 0, "skipped": 0}
            cats[cat]["total"] += 1
            s = tr.get("status", "error")
            if s == "passed":
                cats[cat]["passed"] += 1
            elif s == "failed":
                cats[cat]["failed"] += 1
            elif s == "skipped":
                cats[cat]["skipped"] += 1
            else:
                cats[cat]["errors"] += 1
        return cats


report_generator = ReportGenerator()